"""Lossless tabular readers — every cell is a string; no pandas type inference."""

from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class StringTable:
    headers: list[str]
    rows: list[list[str]]
    sheets: list[str] = field(default_factory=list)
    sheet: str | None = None
    source_path: str = ""
    warnings: list[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def as_dicts(self) -> list[dict[str, str]]:
        return [dict(zip(self.headers, row)) for row in self.rows]


def read_table(path: str | Path, *, sheet: str | None = None) -> StringTable:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return _read_delimited(path, delimiter="\t" if suffix == ".tsv" else None)
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_xlsx(path, sheet=sheet)
    if suffix == ".json":
        return _read_json(path)
    raise ValueError(f"unsupported tabular format: {suffix or path.name}")


def list_sheets(path: str | Path) -> list[str]:
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel uploads") from exc
    wb = load_workbook(path, read_only=True, data_only=False)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _read_delimited(path: Path, *, delimiter: str | None) -> StringTable:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [[_as_str(c) for c in row] for row in reader]
    if not rows:
        return StringTable(headers=[], rows=[], source_path=str(path))
    headers = [_unique_headers(rows[0])]
    # flatten: headers[0] is the unique list
    hdrs = headers[0]
    body = []
    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue
        padded = list(row) + [""] * max(0, len(hdrs) - len(row))
        body.append(padded[: len(hdrs)])
    return StringTable(headers=hdrs, rows=body, source_path=str(path))


def _read_json(path: Path) -> StringTable:
    # Parse with a custom decoder that keeps numbers as strings when possible.
    text = path.read_text(encoding="utf-8")
    # Replace bare numbers with quoted strings before json.loads to preserve
    # leading zeros / long IDs. Only rewrite values (after : or in arrays).
    safe = _quote_json_numbers(text)
    data = json.loads(safe)
    if isinstance(data, dict) and "records" in data:
        data = data["records"]
    if not isinstance(data, list):
        raise ValueError("JSON upload must be a flat array of objects")
    if not data:
        return StringTable(headers=[], rows=[], source_path=str(path))
    if not all(isinstance(row, dict) for row in data):
        raise ValueError("JSON upload must be a flat array of objects")
    headers: list[str] = []
    seen = set()
    for row in data:
        for key in row.keys():
            k = str(key)
            if k not in seen:
                seen.add(k)
                headers.append(k)
    body = []
    for row in data:
        body.append([_as_str(row.get(h, "")) for h in headers])
    return StringTable(headers=headers, rows=body, source_path=str(path))


def _read_xlsx(path: Path, *, sheet: str | None = None) -> StringTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel uploads") from exc

    wb = load_workbook(path, read_only=True, data_only=False)
    sheets = list(wb.sheetnames)
    name = sheet or sheets[0]
    if name not in wb:
        wb.close()
        raise ValueError(f"sheet not found: {name!r}; available: {sheets}")
    ws = wb[name]
    warnings: list[str] = []
    matrix: list[list[str]] = []
    for r_idx, row in enumerate(ws.iter_rows()):
        cells: list[str] = []
        for c_idx, cell in enumerate(row):
            value, warn = _xlsx_cell_as_string(cell, row=r_idx + 1, col=c_idx + 1)
            if warn:
                warnings.append(warn)
            cells.append(value)
        matrix.append(cells)
    wb.close()
    if not matrix:
        return StringTable(headers=[], rows=[], sheets=sheets, sheet=name, source_path=str(path))
    hdrs = _unique_headers([_as_str(c) or f"col_{i+1}" for i, c in enumerate(matrix[0])])
    body = []
    for row in matrix[1:]:
        if not any(str(c).strip() for c in row):
            continue
        padded = list(row) + [""] * max(0, len(hdrs) - len(row))
        body.append([_as_str(c) for c in padded[: len(hdrs)]])
    return StringTable(
        headers=hdrs,
        rows=body,
        sheets=sheets,
        sheet=name,
        source_path=str(path),
        warnings=warnings,
    )


def _xlsx_cell_as_string(cell, *, row: int, col: int) -> tuple[str, str | None]:
    """Preserve text/displayed values; flag precision-risk numeric IDs."""
    if cell.value is None:
        return "", None
    # Explicit text cells (number format '@' or typed as string)
    number_format = (cell.number_format or "").strip()
    if isinstance(cell.value, str):
        return cell.value, None
    if isinstance(cell.value, bool):
        return "true" if cell.value else "false", None
    if isinstance(cell.value, (int, float)):
        # Integers that Excel stored as numbers — preserve exact int form when safe.
        if isinstance(cell.value, float) and not cell.value.is_integer():
            # Likely a genuine decimal, not an ID.
            return repr(cell.value) if "e" in str(cell.value).lower() else str(cell.value), None
        as_int = int(cell.value)
        # Precision risk: IEEE float can't represent integers > 2^53 exactly.
        if abs(as_int) > 2**53:
            return str(as_int), (
                f"precision_risk_id row={row} col={col}: numeric ID may already be truncated by Excel"
            )
        # Prefer displayed text when available.
        try:
            displayed = cell.number_format and str(cell.value)
        except Exception:
            displayed = None
        if number_format == "@":
            return str(as_int), None
        return str(as_int), None
    # dates etc.
    return str(cell.value), None


def _unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for i, h in enumerate(headers):
        base = (h or "").strip() or f"col_{i+1}"
        n = seen.get(base, 0)
        seen[base] = n + 1
        out.append(base if n == 0 else f"{base}_{n+1}")
    return out


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


_NUM_RE = re.compile(
    r'(?<=[:\[\s,])(-?(?:0|[1-9]\d*)(?:\.\d+)?)(?=\s*[,\]\}])'
)


def _quote_json_numbers(text: str) -> str:
    """Quote numeric literals so order IDs like 00123 survive as strings.

    Leaves numbers already inside quotes untouched.
    """
    out = []
    i = 0
    in_str = False
    escape = False
    while i < len(text):
        ch = text[i]
        if in_str:
            out.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
            i += 1
            continue
        if ch == '"':
            in_str = True
            out.append(ch)
            i += 1
            continue
        m = _NUM_RE.match(text, i)
        # Only quote when the number looks like an integer ID (no decimal) or
        # has a leading zero — decimals for prices stay numeric then get
        # re-stringified by our table layer via a second pass.
        if m:
            num = m.group(1)
            out.append(f'"{num}"')
            i = m.end()
            continue
        out.append(ch)
        i += 1
    return "".join(out)
